'''
Created on 27 Apr 2017

@author: friedl
'''

import re
import math

from txttables.tableclass import Table

# package for writing excel tables
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# script providing functions on tables

def readTable(fileName, sep='\t', header=True, headerstart='', colsToRead=None, tableEnd=None, comment=None, split_quoted_cell=True, quote_symbol='"'):
    '''
    function to read a table from a file
    fileName: name of the file to read
    sep: character separating the columns of the table
    header: indicating if the table has a header with column names (first line)
    headerstart: sign at the beginning the header of the table (first line) that will be removed
    colsToRead: a list of indices of column positions to read, giving also their order
    tableEnd: regex for recognizing the end of the table
    comment: sign at linestart indicating a comment -> is not read 
    split_quoted_cell: option for parsing csv files with cells containing ',' symbols which should not be interpreted as cell separators
    quote_symbol: symbol to recognize quoted cells if split_quoted_cell = False
    '''
    
    table = Table()
    
    with open(fileName, 'r') as reader:
        
        # column names in headerS
        if(header==True):
            headerline = reader.readline().strip('\n')
            # remove symbol marking the headerline
            headerline=re.sub('^'+headerstart, '', headerline)
            colNames = _split_row_into_cells(headerline, sep, split_quoted_cell, quote_symbol)#headerline.split(sep)
            
            # read all columns
            if colsToRead is None:
                for colName in colNames:
                    # make all columns of type string
                    table.addColumn(str, colName)
            
            # read specific columns given as column position or column name
            else:
                # create a copy of the list because it gets modified
                colsToRead = colsToRead.copy()
                for (pos,readCol) in enumerate(colsToRead):
                    # column position given -> get column name from header
                    if(isinstance(readCol, int)):
                        table.addColumn(str, colNames[readCol])
                    # column name given -> check with header and extract column position
                    elif(isinstance(readCol, str)):
                        if(readCol in colNames):
                            table.addColumn(str,readCol)
                            colsToRead[pos]=colNames.index(readCol)
                        else:
                            raise ValueError('Cannot read column '+readCol+' because it is not part of the table header!')
                    else:
                        raise ValueError('Cannot read column '+repr(readCol)+' because it is no integer or string!')
                            
                            
        # no header -> create nameless columns
        else:
            # read all columns
            if colsToRead is None:
                firstline = reader.readline().strip('\n')
                while (comment is not None and firstline.startswith(comment)):
                    firstline = reader.readline().strip('\n')
                firstRow = _split_row_into_cells(firstline, sep, split_quoted_cell, quote_symbol) #firstline.split(sep)
                for _ in range(0, len(firstRow)):
                    table.addColumn(str)
                table.addRow(firstRow)
            # read specific columns
            else:
                for _ in range(0, len(colsToRead)):
                    table.addColumn(str)

        
        for line in reader:
            line = line.strip('\n')
            
            # skip comment lines
            if(comment is not None and line.startswith(comment)):
                continue
            
            # check for table end
            if(tableEnd is not None and re.search(tableEnd, line)):
                break
            
            # add current row to table
            rowContent = _split_row_into_cells(line, sep, split_quoted_cell, quote_symbol)#line.split(sep)
            # add all columns of the row
            if(colsToRead is None):
                table.addRow(rowContent)
            # add only specific columns of the row in a given order
            else:
                reducedRowContent=[]
                for pos in colsToRead:
                    reducedRowContent.append(rowContent[pos])
                table.addRow(reducedRowContent)
    
    return table


def _split_row_into_cells(line, separator_symbol, split_quoted, quote_symbol):
    '''
    auxiliary method to split a line of the table into individual cells -> mainly required for handling csv format
    'separator_symbol': symbol that marks the cell borders
    'do_not_split_quote': cells containing the separator_symbol as content are quoted
    '''
    
    # split line into cells every time the separator occurs
    if(split_quoted is True):
        return line.split(separator_symbol)
    
    # split line into cells every time the separator does not occur within quotes
    else:
        res = []
        tempsplit = line.split(separator_symbol)
        quotedregion = False
        quotedcell = ''
        for cell in tempsplit:
            if(quotedregion is False):
                # started of quoted cell: cell starting with a quote symbol but not ending with a quote symbol
                if(cell.startswith(quote_symbol) and not cell.endswith(quote_symbol)):
                    quotedregion = True
                    quotedcell = cell
                # unquoted cell
                else:
                    res.append(cell)
            else:
                quotedcell+=','+cell
                #end of quoted cell
                if(cell.endswith(quote_symbol)):
                    res.append(quotedcell)
                    quotedcell=''
                    quotedregion = False
        return res



def writeTable(table, fileName, sep='\t', header=True, headerstart='', colsToWrite=None):
    '''
    write table to file
    '''
    
    with open(fileName, 'w') as tabWriter:
        
        # construct table header
        if(header==True):
            fileHead= _get_table_header_output(table, colsToWrite)
            tabWriter.write(headerstart+sep.join(fileHead)+'\n')
        
        # iterate over all rows, extract the desired values and build a line for the file
        for rowInd in range(0, table.rowNum()):
            tabWriter.write(sep.join(map(str,table.getRow(rowInd, select_cols=colsToWrite)))+'\n')
            

def writeExcelTable(table, fileName, header=True, colsToWrite=None, adapt_column_widths=False):
    '''
    write table to xls file
    '''
    
    wb = Workbook()
    ws = wb.active
    
    if header is True:
        headerRow = _get_table_header_output(table, colsToWrite)
        ws.append(headerRow)
    
    # iterate over all rows, extract the desired columns of the current row and add them to the worksheet
    for rowInd in range(0, table.rowNum()):
        ws.append(table.getRow(rowInd, select_cols=colsToWrite))
        
    if(adapt_column_widths):
        if colsToWrite is None:
            cols_to_adapt = range(0, table.colNum())
        else:
            cols_to_adapt = colsToWrite
        for p,c in enumerate(cols_to_adapt):
            lens = [len(str(val)) for val in table.getColumn(c)]
            if(header is True):
                lens.append(len(table.getColumnName(c)))
            ws.column_dimensions[get_column_letter(p+1)].width = max(lens)*1.3     
    
    wb.save(fileName)
    

def _get_table_header_output(table, colsToWrite):
    '''
    function used for writing tables to file that constructs the table header
    '''
    
    # list of column names to write as table header
    fileHead=[]
    
    # colsToWrite = None -> write all columns
    colnames = table.getColumnNames()
    if(colsToWrite is None):
        fileHead=colnames
        
    # build custom header, if not all columns are written
    else:
        for outCol in colsToWrite:
            if(isinstance(outCol, int)):
                if(outCol<len(colnames)):
                    fileHead.append(colnames[outCol])
                else:
                    raise ValueError('Cannot write table: no column at position '+str(outCol))
            else:
                if(outCol in colnames):
                    fileHead.append(outCol)
                else:
                    raise ValueError('Cannot write table: no column named '+str(outCol))
    
    # return list of all column names to write into the table header
    return fileHead


def selectRows(table, selectionFunction):
    '''
    method to create a new table with a subset of rows from 'table'
    selectionFunction: input table+rowindex -> output: true/false, true <-> select column for new table 
    '''
    
    # create new table with same columns as 'table'
    selection_table = Table(table)
    
    # iterate over all rows of 'table' and add them to the new table if they meet the selection criteria
    for rowPos in range(0,table.rowNum()):
        if(selectionFunction(table, rowPos)):
            selection_table.addRow(table.getRow(rowPos))
    
    return selection_table


def selectColumns(table, colList):
    '''
    creates a new table with all columns in colList in the given order
    '''
    
    # create new table
    selection_table = Table()
    
    # create columns of new table
    for column in colList:
        selection_table.addColumn(table.getColumnType(column), table.getColumnName(column))            
    
    # iterate over all rows of 'table' and extract the data for the columns of 'colList'
    for rowPos in range(0, table.rowNum()):
        selection_table.addRow(table.getRow(rowPos, select_cols=colList))
    
    return selection_table

    
def joinTables(tab1, tab2, joinCols, joinType='inner'):
    '''
    join 2 tables tab1 and tab2
    joinCols: list of tuples (col_1, col_2) to compare when joining, col_i: column position or name from tab_i,
    join type: 'inner', 'leftouter' and 'fullouter'
    '''
    
    # TODO: check types of columns
    
    # TODO: implement outer join types: rightouter
    
        
    # set up configuration of joined table -> column names and types combined
    
    jcols1, jcols2 = zip(*joinCols)
    joined_table = Table()
    
    # take all columns of the first table
    for colPos, colName in enumerate(tab1.getColumnNames(noneHandling=False)):
        if(colName is None):
            joined_table.addColumn(tab1.getColumnType(colPos))
        else:
            joined_table.addColumn(tab1.getColumnType(colPos), colName)
            
    # take all columns except the jcols2
    jcol_set = set(jcols2)
    # indices of columns to keep
    cols_keep=[]
    for colPos, colName in enumerate(tab2.getColumnNames()):
        if((not colPos in jcol_set) and (not colName in jcol_set)):
            if(colName is None):
                joined_table.addColumn(tab2.getColumnType(colPos))
            else:
                joined_table.addColumn(tab2.getColumnType(colPos), colName)
            cols_keep.append(colPos)
            
    # build indices mapping values of joinCols to row index positions for each table
    
    # index = mapping jcol2 values to row indices in tab2
    index={}
    for i in range(0, tab2.rowNum()):
        keyInd = tuple(map(lambda c: tab2.get(i,c), jcols2))
        if(not keyInd in index):
            index[keyInd] = [i]
        else:
            index[keyInd].append(i)
    # seenKeys = all jcol2 values with a join partner in table1 -> use for full outer join
    seenKeys = set()
    
    # join rows and fill joined table
    
    # iterate over table 1 and and join rows
    for i in range(0, tab1.rowNum()):
        index_key = tuple(map(lambda c: tab1.get(i,c), jcols1))
        # inner join: all pairs of rows with matching join-condition
        if(index_key in index):
            for row2Ind in index[index_key]:
                newRow = tab1.getRow(i)
                for pos in cols_keep:
                    newRow.append(tab2.get(row2Ind, pos))
                joined_table.addRow(newRow)
            seenKeys.add(index_key)
        
        # left outer join & full outer join: keep rows from table 1 without join partner
        elif(joinType=='leftouter' or joinType=='fullouter'):
            newRow=tab1.getRow(i)
            for _ in cols_keep:
                newRow.append(None)
            joined_table.addRow(newRow)
    
    # full outer join: add rows from table 2 without join partner
    if(joinType=='fullouter'):
        for index_key in index.keys():
            if not index_key in seenKeys:
                for rowInd in index[index_key]:
                    newRow=[]
                    # no join partner from table 1 -> set columns from table 1 to None
                    for _ in range(0,tab1.colNum()):
                        newRow.append(None)
                    # join columns only once in output -> set values from table2 join Columns for table1 join Columns
                    for keyCol1, keyCol2 in joinCols:
                        keyCol1_pos = tab1._check_table_access_column(keyCol1, 'join')
                        newRow[keyCol1_pos]=tab2.get(rowInd, keyCol2)
                    # add data from table 2
                    for col in cols_keep:
                        newRow.append(tab2.get(rowInd, col))
                    joined_table.addRow(newRow)
    
    return joined_table

def joinTableList(table_list, joinCols, joinType):
    '''
    joins all tables from the list, assumes all tables of the list share the join columns (names or positions)
    joinCols: list of column names or positions
    join type: 'inner', 'leftouter' and 'fullouter'
    '''
    
    res_table = table_list[0].copy()
    join_tuples = [(x,x) for x in joinCols]
    for next_table in table_list[1:]:
        res_table = joinTables(res_table, next_table, joinCols=join_tuples, joinType=joinType)
    return res_table
    
# TODO: join function with function as join criterion

#TODO: implementation of additional aggregation functions, 
def groupBy(table, groupColumns, aggregateColumns, aggregateFunctions, append_aggregation_func_to_colname=False):
    '''
    'groupColumns': list of column names or indices used for forming groups, aggregates all rows if set to []
    'aggregateColumns': list of column names that are aggregated into a single value per column and group
    'aggregateFunctions': list of strings giving aggregation functions
    'append_aggregation_func_to_colname': defines if column names of aggregate columns are modified in the returned table
    '''            
    
    ## check input
    
    # check if groupColumns and aggregateColumns are passed as single element -> transform into iterables
    if(isinstance(groupColumns, int) or isinstance(groupColumns, str)):
        groupColumns=[groupColumns]
    if(isinstance(aggregateColumns, int) or isinstance(aggregateColumns, str)):
        aggregateColumns=[aggregateColumns]
    if(isinstance(aggregateFunctions, str)):
        aggregateFunctions=[aggregateFunctions]
    
    # transform aggregateFunctions given as strings to function pointers
    aggregate_fps=[]
    for af_name in aggregateFunctions:
        if(af_name=='min'):
            aggregate_fps.append(_min)
        elif(af_name=='sum'):
            aggregate_fps.append(_sum)
        else:
            raise ValueError('Unknown aggregation function: '+str(af_name))
    
    ## mapping of groups to row indices
    
    # map group as tuple of values -> list of row indices as 0-based ints
    group_to_rows={}
    for rowInd in range(0, table.rowNum()):
        
        # get group of the current row
        group=[]
        for col in groupColumns:
            group.append(table.get(rowInd, col))
        group = tuple(group)
        
        # add row to the mapping
        if(not group in group_to_rows):
            group_to_rows[group]=[rowInd]
        else:
            group_to_rows[group].append(rowInd)
    
    ## set up return table     
       
    groupTable = Table()
    # add columns defining the groups
    for gc in groupColumns:
        groupTable.addColumn(table.getColumnType(gc), table.getColumnName(gc))
    # add columns for aggregated values
    for ac,af in zip(aggregateColumns, aggregateFunctions):
        if(append_aggregation_func_to_colname is False):
            groupTable.addColumn(table.getColumnType(ac), table.getColumnName(ac))
        else:
            groupTable.addColumn(table.getColumnType(ac), table.getColmnName(ac)+'_'+af)
    
    ##  perform aggregations 
    
    for cur_group,cur_rowlist in sorted(group_to_rows.items()):
        row = []
        for group_element in cur_group:
            row.append(group_element)
        for ac, af in zip(aggregateColumns, aggregate_fps):
            aggregated_value = af(table, cur_rowlist, ac)
            row.append(aggregated_value)
        groupTable.addRow(row)
    
    return groupTable
    
    
def _min(table, row_index_list, column):
    '''
    aggregation function -> minimum of a column for the rows given in index_list
    '''
    
    curmin =None
    if(table.getColumnType(column) == float or table.getColumnType(column) == int):
        curmin = float('nan')
    else:
        raise ValueError('Aggregation Function Minimum does not support non-numeric column types: '+str(table.getColumnType(column)))
    
    for row in row_index_list:
        col_value = table.get(row, column)
        # check if current number is defined
        if(not math.isnan(col_value)):
            # update minimum
            if(math.isnan(curmin) or col_value<curmin):
                curmin=col_value
    
    return curmin

def _sum(table, row_index_list, column):
    '''
    aggregation function -> sum of a column for the rows given in index_list
    '''
    
    cursum=None
    if(table.getColumnType(column) == float or table.getColumnType(column) == int):
        cursum = 0
    else:
        raise ValueError('Aggregation Function Sum does not support non-numeric column types: '+str(table.getColumnType(column)))
    
    for row in row_index_list:
        col_value = table.get(row, column)
        # check if current number is defined
        if(not math.isnan(col_value)):
            # update minimum
            cursum+=col_value
    
    return cursum
    